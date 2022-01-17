import sys
import webbrowser

import openpyxl
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from urllib.request import urlopen
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

class moje_okno(QMainWindow):
    def __init__(self):
        super(moje_okno,self).__init__()
        self.setGeometry(500,500,300,220)
        self.setWindowTitle("Projekt 2.0")

        #PRZYCISKI
        self.p_kupno = QPushButton(self)
        self.p_kupno.setText("Kupno")
        self.p_kupno.setGeometry(50,10,200,50)
        self.kupno_okno = Kupno()
        self.p_kupno.clicked.connect(self.kupno_okno.show)

        self.p_sprzed = QPushButton(self)
        self.p_sprzed.setText("Sprzedaż")
        self.p_sprzed.setGeometry(50,60, 200, 50)
        self.sprzed_okno = Sprzedaz()
        self.p_sprzed.clicked.connect(self.sprzed_okno.show)

        self.p_aktualnosci = QPushButton(self)
        self.p_aktualnosci.setText("Aktualności")
        self.p_aktualnosci.setGeometry(50,110, 200, 50)
        self.aktual_okno = Aktualnosci()
        self.p_aktualnosci.clicked.connect(self.aktual_okno.show)

        self.p_kurs = QPushButton(self)
        self.p_kurs.setText("Aktualny kurs walut")
        self.p_kurs.setGeometry(50,160, 200, 50)
        self.kurs_okno = Aktualny_kurs()
        self.p_kurs.clicked.connect(self.kurs_okno.show)



class Kupno(QWidget):
    def __init__(self):
        super(Kupno, self).__init__()
        self.setGeometry(500,500,750,360)
        self.setWindowTitle("Kupno")

        self.lista1 = QListWidget(self)
        self.lista1.setGeometry(20,10,200,80)
        self.lista1.addItems(["Euro", "Dolar Amrykański","Funt Szwajcarski","Funt Brytyjski"])
        self.lista1.setCurrentRow(0)


        self.przelicz = QPushButton(self)
        self.przelicz.setGeometry(20,300,400,40)
        self.przelicz.setText("&Przelicz")
        self.przelicz.clicked.connect(self.przeliczenie)

        self.pole_tekst = QLineEdit(self)
        self.pole_tekst.setGeometry(50,220,130,50)
        self.pole_tekst.setFont(QFont('Comic Sans',16))

        self.wynik = QLabel(self)
        self.wynik.setGeometry(230,60,200,100)
        self.wynik.setFont(QFont('Arial',16))

        self.waluta = QLabel(self)
        self.waluta.setGeometry(330, 60, 90, 100)
        self.waluta.setFont(QFont('Arial', 16))

        self.tabela = QTableWidget(self)
        self.tabela.setGeometry(420, 10, 320, 340)
        self.tabela.setColumnCount(4)
        self.tabela.setColumnWidth(0, 100)
        self.tabela.setColumnWidth(1, 70)
        self.tabela.setColumnWidth(2, 70)
        self.tabela.setColumnWidth(3, 70)
        self.tabela.setHorizontalHeaderLabels(("Zamiana", "Kurs", "Ilość","Wartość"))
        self.tabela.verticalHeader().hide()
        wb = load_workbook('Kantor.xlsx')

        sheet = wb['Arkusz1']
        for i in range(1,30):
            Zamiana = sheet.cell(row=i, column=1).value
            Kurs = sheet.cell(row=i, column=2).value
            Ilość = sheet.cell(row=i, column=3).value
            Wartość = sheet.cell(row=i, column=4).value
            if Zamiana == None:
                break
            else:
                obecny_rząd = self.tabela.rowCount()
                self.tabela.insertRow(obecny_rząd)
                self.tabela.setItem(obecny_rząd, 0, QTableWidgetItem(Zamiana))
                self.tabela.setItem(obecny_rząd, 1, QTableWidgetItem(Kurs))
                self.tabela.setItem(obecny_rząd, 2, QTableWidgetItem(Ilość))
                self.tabela.setItem(obecny_rząd, 3, QTableWidgetItem(Wartość))




    @pyqtSlot()
    def przeliczenie(self):
        nadawca = self.sender()

        html = urlopen('https://kantoronline.pl/kursy-walut')
        bs = BeautifulSoup(html.read(), 'html.parser')
        eurkp = bs.find(id='eurbuy')
        euro_kup = eurkp.getText()
        usdkp = bs.find(id='usdbuy')
        usd_kup = usdkp.getText()
        chfkp = bs.find(id='chfbuy')
        chf_kup = chfkp.getText()
        gbpkp = bs.find(id='gbpbuy')
        gbp_kup = gbpkp.getText()

        obecny_rząd = self.tabela.rowCount()
        self.tabela.insertRow(obecny_rząd)
        plneur = "PLN->EUR"
        plnusd = "PLN->USD"
        plnchf = "PLN->CHF"
        plngbp = "PLN->GBP"

        wb = openpyxl.load_workbook('Kantor.xlsx')
        sheet = wb.active
        sheet.title = 'Arkusz1'


        try:

            euro = self.lista1.currentRow() == 0
            dolar = self.lista1.currentRow() == 1
            szwajcar = self.lista1.currentRow() == 2
            bryt = self.lista1.currentRow() == 3
            wartosc1 = float(self.pole_tekst.text())
            wartosc2 = float(euro_kup)
            wartosc3 = float(usd_kup)
            wartosc4 = float(chf_kup)
            wartosc5 = float(gbp_kup)
            przel = ""

            if euro and nadawca.text() == "&Przelicz":
                przel = wartosc1 / wartosc2
                wwynik = round(przel,2)
                self.tabela.setItem(obecny_rząd, 0,QTableWidgetItem(plneur))
                self.tabela.setItem(obecny_rząd, 1, QTableWidgetItem(euro_kup))
                self.tabela.setItem(obecny_rząd, 2, QTableWidgetItem(self.pole_tekst.text()))
                sheet.append([plneur,euro_kup,self.pole_tekst.text(),str(wwynik)])
                self.waluta.setText('EUR')

            elif dolar and nadawca.text() == "&Przelicz":
                przel = wartosc1 / wartosc3
                wwynik = round(przel,2)
                self.tabela.setItem(obecny_rząd, 0,QTableWidgetItem(plnusd))
                self.tabela.setItem(obecny_rząd, 1, QTableWidgetItem(usd_kup))
                self.tabela.setItem(obecny_rząd, 2, QTableWidgetItem(self.pole_tekst.text()))
                sheet.append([plnusd,usd_kup,self.pole_tekst.text(),str(wwynik)])
                self.waluta.setText('USD')

            elif szwajcar and nadawca.text() == "&Przelicz":
                przel = wartosc1 / wartosc4
                wwynik = round(przel,2)
                self.tabela.setItem(obecny_rząd, 0,QTableWidgetItem(plnchf))
                self.tabela.setItem(obecny_rząd, 1, QTableWidgetItem(chf_kup))
                self.tabela.setItem(obecny_rząd, 2, QTableWidgetItem(self.pole_tekst.text()))
                sheet.append([plnchf,chf_kup,self.pole_tekst.text(),str(wwynik)])
                self.waluta.setText('CHF')

            elif bryt and nadawca.text() == "&Przelicz":
                przel = wartosc1 / wartosc5
                wwynik = round(przel,2)
                self.tabela.setItem(obecny_rząd, 0,QTableWidgetItem(plngbp))
                self.tabela.setItem(obecny_rząd, 1, QTableWidgetItem(gbp_kup))
                self.tabela.setItem(obecny_rząd, 2, QTableWidgetItem(self.pole_tekst.text()))
                sheet.append([plngbp,gbp_kup,self.pole_tekst.text(),str(wwynik)])
                self.waluta.setText('GBP')

            else:
                QMessageBox.critical(
                    self, "Błąd", "Nie wybrano waluty!")
                return

            self.wynik.setText(str(wwynik))
            self.tabela.setItem(obecny_rząd, 3, QTableWidgetItem(self.wynik.text()))
            wb.save('Kantor.xlsx')

        except ValueError:
            QMessageBox.warning(self, "Błąd", "Coś poszło nie tak...", QMessageBox.Ok)




class Sprzedaz(QWidget):
    def __init__(self):
        super(Sprzedaz, self).__init__()
        self.setGeometry(500, 500, 750, 360)
        self.setWindowTitle("Sprzedaż")

        self.lista1 = QListWidget(self)
        self.lista1.setGeometry(20, 10, 200, 80)
        self.lista1.addItems(["Euro", "Dolar Amrykański", "Funt Szwajcarski", "Funt Brytyjski"])
        self.lista1.setCurrentRow(0)

        self.przelicz = QPushButton(self)
        self.przelicz.setGeometry(20, 300, 400, 40)
        self.przelicz.setText("&Przelicz")
        self.przelicz.clicked.connect(self.przeliczenie)

        self.pole_tekst = QLineEdit(self)
        self.pole_tekst.setGeometry(50, 220, 130, 50)
        self.pole_tekst.setFont(QFont('Comic Sans', 16))

        self.wynik = QLabel(self)
        self.wynik.setGeometry(230, 60, 200, 100)
        self.wynik.setFont(QFont('Arial', 16))

        self.waluta = QLabel(self)
        self.waluta.setGeometry(330, 60, 90, 100)
        self.waluta.setFont(QFont('Arial', 16))

        self.tabela = QTableWidget(self)
        self.tabela.setGeometry(420, 10, 320, 340)
        self.tabela.setColumnCount(4)
        self.tabela.setColumnWidth(0, 100)
        self.tabela.setColumnWidth(1, 70)
        self.tabela.setColumnWidth(2, 70)
        self.tabela.setColumnWidth(3, 70)
        self.tabela.setHorizontalHeaderLabels(("Zamiana", "Kurs", "Ilość", "Wartość"))
        self.tabela.verticalHeader().hide()
        wb = load_workbook('Kantor.xlsx')

        sheet = wb['Arkusz1']
        for i in range(1, 30):
            Zamiana = sheet.cell(row=i, column=1).value
            Kurs = sheet.cell(row=i, column=2).value
            Ilość = sheet.cell(row=i, column=3).value
            Wartość = sheet.cell(row=i, column=4).value
            if Zamiana == None:
                break
            else:
                obecny_rząd = self.tabela.rowCount()
                self.tabela.insertRow(obecny_rząd)
                self.tabela.setItem(obecny_rząd, 0, QTableWidgetItem(Zamiana))
                self.tabela.setItem(obecny_rząd, 1, QTableWidgetItem(Kurs))
                self.tabela.setItem(obecny_rząd, 2, QTableWidgetItem(Ilość))
                self.tabela.setItem(obecny_rząd, 3, QTableWidgetItem(Wartość))

    @pyqtSlot()
    def przeliczenie(self):
        nadawca = self.sender()

        html = urlopen('https://kantoronline.pl/kursy-walut')
        bs = BeautifulSoup(html.read(), 'html.parser')
        eurkp = bs.find(id='eursell')
        euro_kup = eurkp.getText()
        usdkp = bs.find(id='usdsell')
        usd_kup = usdkp.getText()
        chfkp = bs.find(id='chfsell')
        chf_kup = chfkp.getText()
        gbpkp = bs.find(id='gbpsell')
        gbp_kup = gbpkp.getText()

        obecny_rząd = self.tabela.rowCount()
        self.tabela.insertRow(obecny_rząd)
        eurpln = "EUR->PLN"
        usdpln = "USD->PLN"
        chfpln = "CHF->PLN"
        gbppln = "GBP->PLN"

        wb = openpyxl.load_workbook('Kantor.xlsx')
        sheet = wb.active
        sheet.title = 'Arkusz1'

        try:

            euro = self.lista1.currentRow() == 0
            dolar = self.lista1.currentRow() == 1
            funt = self.lista1.currentRow() == 2
            peso = self.lista1.currentRow() == 3
            wartosc1 = float(self.pole_tekst.text())
            wartosc2 = float(euro_kup)
            wartosc3 = float(usd_kup)
            wartosc4 = float(chf_kup)
            wartosc5 = float(gbp_kup)
            przel = ""

            if euro and nadawca.text() == "&Przelicz":
                przel = wartosc1 * wartosc2
                wwynik = round(przel,2)
                self.tabela.setItem(obecny_rząd, 0, QTableWidgetItem(eurpln))
                self.tabela.setItem(obecny_rząd, 1, QTableWidgetItem(euro_kup))
                self.tabela.setItem(obecny_rząd, 2, QTableWidgetItem(self.pole_tekst.text()))
                sheet.append([eurpln, euro_kup, self.pole_tekst.text(), str(wwynik)])

            elif dolar and nadawca.text() == "&Przelicz":
                przel = wartosc1 * wartosc3
                wwynik = round(przel, 2)
                self.tabela.setItem(obecny_rząd, 0, QTableWidgetItem(usdpln))
                self.tabela.setItem(obecny_rząd, 1, QTableWidgetItem(usd_kup))
                self.tabela.setItem(obecny_rząd, 2, QTableWidgetItem(self.pole_tekst.text()))
                sheet.append([usdpln, usd_kup, self.pole_tekst.text(), str(wwynik)])

            elif funt and nadawca.text() == "&Przelicz":
                przel = wartosc1 * wartosc4
                wwynik = round(przel, 2)
                self.tabela.setItem(obecny_rząd, 0, QTableWidgetItem(chfpln))
                self.tabela.setItem(obecny_rząd, 1, QTableWidgetItem(chf_kup))
                self.tabela.setItem(obecny_rząd, 2, QTableWidgetItem(self.pole_tekst.text()))
                sheet.append([chfpln, chf_kup, self.pole_tekst.text(), str(wwynik)])

            elif peso and nadawca.text() == "&Przelicz":
                przel = wartosc1 * wartosc5
                wwynik = round(przel, 2)
                self.tabela.setItem(obecny_rząd, 0, QTableWidgetItem(gbppln))
                self.tabela.setItem(obecny_rząd, 1, QTableWidgetItem(gbp_kup))
                self.tabela.setItem(obecny_rząd, 2, QTableWidgetItem(self.pole_tekst.text()))
                sheet.append([gbppln, gbp_kup, self.pole_tekst.text(), str(wwynik)])

            else:
                QMessageBox.critical(
                    self, "Błąd", "Nie wybrano waluty!")
                return

            self.wynik.setText(str(wwynik))
            self.waluta.setText('PLN')
            self.tabela.setItem(obecny_rząd, 3, QTableWidgetItem(self.wynik.text()))
            wb.save('Kantor.xlsx')

        except ValueError:
            QMessageBox.warning(self, "Błąd", "Coś poszło nie tak...", QMessageBox.Ok)

class Aktualnosci(QWidget):
    def __init__(self):
        html = urlopen('https://www.pb.pl/puls-inwestora/waluty/')
        bs = BeautifulSoup(html.read(), 'html.parser')
        tytul = bs.find_all('h2')
        podglad = bs.find_all('p')

        super(Aktualnosci, self).__init__()
        self.setGeometry(500, 500, 740, 320)
        self.setWindowTitle("Aktualnosci")

        self.podglad = QTextEdit(self)
        self.podglad.setGeometry(10, 10, 440, 230)
        self.podglad.setReadOnly(True)

        self.przycisk = QPushButton(self)
        self.przycisk.setGeometry(10, 250, 720, 50)
        self.przycisk.setText("&Przejdź do pełnego artykułu")
        self.przycisk.clicked.connect(self.przycisnij)

        self.lista = QListWidget(self)
        self.lista.setGeometry(470, 30, 260, 190)
        self.lista.setCurrentRow(0)
        self.lista.itemDoubleClicked.connect(self.pokaz)



        for x in tytul:
            self.lista.addItem(x.get_text())
        self.lista.setRowHidden(15, True)
        self.lista.setRowHidden(16, True)
        self.lista.setRowHidden(17, True)

    @pyqtSlot()
    def przycisnij(self):
        nadawca = self.sender()
        html = urlopen('https://www.pb.pl/puls-inwestora/waluty/')
        bs = BeautifulSoup(html.read(), 'html.parser')
        podglad = bs.find_all('p')
        linki = []
        for link in bs.find_all('a'):
            linki.append(link.get('href'))

        t1 = self.lista.currentRow() == 0
        t2 = self.lista.currentRow() == 1
        t3 = self.lista.currentRow() == 2
        t4 = self.lista.currentRow() == 3
        t5 = self.lista.currentRow() == 4
        t6 = self.lista.currentRow() == 5
        t7 = self.lista.currentRow() == 6
        t8 = self.lista.currentRow() == 7
        t9 = self.lista.currentRow() == 8
        t10 = self.lista.currentRow() == 9
        t11 = self.lista.currentRow() == 10
        t12 = self.lista.currentRow() == 11
        t13 = self.lista.currentRow() == 12
        t14 = self.lista.currentRow() == 13
        t15 = self.lista.currentRow() == 14

        if t1 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[97])
            webbrowser.open_new_tab(artlink)
        elif t2 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[98])
            webbrowser.open_new_tab(artlink)
        elif t3 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[99])
            webbrowser.open_new_tab(artlink)
        elif t4 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[100])
            webbrowser.open_new_tab(artlink)
        elif t5 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[101])
            webbrowser.open_new_tab(artlink)
        elif t6 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[102])
            webbrowser.open_new_tab(artlink)
        elif t7 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[103])
            webbrowser.open_new_tab(artlink)
        elif t8 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[104])
            webbrowser.open_new_tab(artlink)
        elif t9 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[105])
            webbrowser.open_new_tab(artlink)
        elif t10 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[106])
            webbrowser.open_new_tab(artlink)
        elif t11 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[107])
            webbrowser.open_new_tab(artlink)
        elif t12 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[108])
            webbrowser.open_new_tab(artlink)
        elif t13 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[109])
            webbrowser.open_new_tab(artlink)
        elif t14 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[110])
            webbrowser.open_new_tab(artlink)
        elif t15 and nadawca.text() == "&Przejdź do pełnego artykułu":
            artlink = ("https://pb.pl" + linki[111])
            webbrowser.open_new_tab(artlink)


    @pyqtSlot()
    def pokaz(self):
        html = urlopen('https://www.pb.pl/puls-inwestora/waluty/')
        bs = BeautifulSoup(html.read(), 'html.parser')
        podglad = bs.find_all('p')
        linki = []
        for link in bs.find_all('a'):
            linki.append(link.get('href'))

        t1 = self.lista.currentRow() == 0
        t2 = self.lista.currentRow() == 1
        t3 = self.lista.currentRow() == 2
        t4 = self.lista.currentRow() == 3
        t5 = self.lista.currentRow() == 4
        t6 = self.lista.currentRow() == 5
        t7 = self.lista.currentRow() == 6
        t8 = self.lista.currentRow() == 7
        t9 = self.lista.currentRow() == 8
        t10 = self.lista.currentRow() == 9
        t11 = self.lista.currentRow() == 10
        t12 = self.lista.currentRow() == 11
        t13 = self.lista.currentRow() == 12
        t14 = self.lista.currentRow() == 13
        t15 = self.lista.currentRow() == 14

        if t1:
            self.podglad.setText(podglad[0].get_text())
        elif t2:
            self.podglad.setText(podglad[1].get_text())
        elif t3:
            self.podglad.setText(podglad[2].get_text())
        elif t4:
            self.podglad.setText(podglad[3].get_text())
        elif t5:
            self.podglad.setText(podglad[4].get_text())
        elif t6:
            self.podglad.setText(podglad[5].get_text())
        elif t7:
            self.podglad.setText(podglad[6].get_text())
        elif t8:
            self.podglad.setText(podglad[7].get_text())
        elif t9:
            self.podglad.setText(podglad[8].get_text())
        elif t10:
            self.podglad.setText(podglad[9].get_text())
        elif t11:
            self.podglad.setText(podglad[10].get_text())
        elif t12:
            self.podglad.setText(podglad[11].get_text())
        elif t13:
            self.podglad.setText(podglad[12].get_text())
        elif t14:
            self.podglad.setText(podglad[13].get_text())
        elif t15:
            self.podglad.setText(podglad[14].get_text())





            



class Aktualny_kurs(QWidget):
    def __init__(self):
        super(Aktualny_kurs, self).__init__()
        self.setGeometry(500,500,480,150)
        self.setWindowTitle("Aktualny kurs")

        self.tabela = QTableWidget(self)
        self.tabela.setRowCount(4)
        self.tabela.setColumnCount(3)
        self.tabela.setColumnWidth(0,120)
        self.tabela.setGeometry(0,0,380,150)
        self.tabela.setHorizontalHeaderLabels(("Waluta","Kupno","Sprzedaż"))
        self.tabela.verticalHeader().hide()
        self.tabela.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabela.setItem(0,0,QTableWidgetItem("Euro"))
        self.tabela.setItem(1,0,QTableWidgetItem("Dolar Amerykański"))
        self.tabela.setItem(2,0,QTableWidgetItem("Funt Szwajcarski"))
        self.tabela.setItem(3,0,QTableWidgetItem("Funt Brytyjski"))

        self.przycisk = QPushButton(self)
        self.przycisk.setGeometry(380, 0, 100, 150)
        self.przycisk.setText("&Odśwież")
        self.przycisk.clicked.connect(self.odswiez)

    @pyqtSlot()
    def odswiez(self):
        nadawca = self.sender()
        html = urlopen('https://kantoronline.pl/kursy-walut')
        bs = BeautifulSoup(html.read(), 'html.parser')
        eurkp = bs.find(id='eursell')
        euro_kup = eurkp.getText()
        usdkp = bs.find(id='usdsell')
        usd_kup = usdkp.getText()
        chfkp = bs.find(id='chfsell')
        chf_kup = chfkp.getText()
        gbpkp = bs.find(id='gbpsell')
        gbp_kup = gbpkp.getText()
        eurs = bs.find(id='eurbuy')
        euro_s = eurs.getText()
        usds = bs.find(id='usdbuy')
        usd_s = usds.getText()
        chfs = bs.find(id='chfbuy')
        chf_s = chfs.getText()
        gbps = bs.find(id='gbpbuy')
        gbp_s = gbps.getText()

        if nadawca.text() == "&Odśwież":
            self.tabela.setItem(0, 1, QTableWidgetItem(euro_kup))
            self.tabela.setItem(1, 1, QTableWidgetItem(usd_kup))
            self.tabela.setItem(2, 1, QTableWidgetItem(chf_kup))
            self.tabela.setItem(3, 1, QTableWidgetItem(gbp_kup))
            self.tabela.setItem(0, 2, QTableWidgetItem(euro_s))
            self.tabela.setItem(1, 2, QTableWidgetItem(usd_s))
            self.tabela.setItem(2, 2, QTableWidgetItem(chf_s))
            self.tabela.setItem(3, 2, QTableWidgetItem(gbp_s))
        else:
            QMessageBox.warning(self, "Błąd", "Brak danych", QMessageBox.Ok)








if __name__ == '__main__':
    app = QApplication([])
    window = moje_okno()
    window.show()
    sys.exit(app.exec())